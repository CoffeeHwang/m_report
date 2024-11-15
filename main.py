import io
import json
import base64
import smtplib
import itertools
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.utils import formataddr
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, PatternFill, Font
from pymysql import connect
import pandas as pd
from copy import copy
from yhs_common import *
from google.cloud import storage, pubsub_v1
import yhs_mysql as yhsdb


def get_ent_name(conn: connect, ent_id: int) -> str:
    sql = "select comment from svc_enterprise where id = %s limit 1"
    return yhsdb.__get_one_yhsdb(conn=conn, sql=sql, args=(ent_id,))

def get_report_recv_emails(conn: connect, ent_id: int) -> str:
    sql = "select value_str from ref_code where grp_cd = 'proc_report_recvs' and cd = %s limit 1"
    return yhsdb.__get_one_yhsdb(conn=conn, sql=sql, args=(ent_id,))

def get_report_ver_hist(conn: connect) -> str:
    sql = "select value_str from ref_code where grp_cd = 'proc_report_ver_hist' and cd = '1' limit 1"
    return yhsdb.__get_one_yhsdb(conn=conn, sql=sql)

def get_report_const_idle_outlier(conn: connect) -> str:
    sql = "select value_num from ref_code where grp_cd = 'proc_report_const' and cd = '1' limit 1"
    return yhsdb.__get_one_yhsdb(conn=conn, sql=sql)

def get_process_report_brktimes(conn: connect, ent_id: int = None) -> pd.DataFrame:
    sql = "select (select t.comment from svc_enterprise t where t.id = a.ent_id) ent_name, a.start_time, a.end_time " \
          "from ref_std_worktime a " \
          "where a.ent_id = %s " \
          "and a.time_mng_type_cd = 'BRK' " \
          "order by a.start_time, a.end_time"
    return yhsdb.__get_yhsdb_query(conn=conn, sql=sql, args=(ent_id,))

def get_process_report(conn: connect, report_date: str, ent_id: int = None) -> pd.DataFrame:
    sql = "select " \
          "`report_date`, " \
          "`ent_id`, " \
          "`mkey`, " \
          "`lot`, " \
          "`집계 대상 업체`, " \
          "`machine_no`, " \
          "`machine_name`, " \
          "`program`, " \
          "`가공개수`, " \
          "`조업 시작`, " \
          "`조업 종료`, " \
          "`조업중 총 휴게시간`, " \
          "`실 조업시간`, " \
          "`실 부하시간`, " \
          "`실 가공시간`, " \
          "`실 조업 대비 비가동시간`, " \
          "`실 부하 대비 비가동시간`, " \
          "`실 조업시간 대비 가동율`, " \
          "`실 부하시간 대비 가동율`, " \
          "`평균 가공시간`, " \
          "`평균 실가공시간`, " \
          "`평균 준비교체시간`, " \
          "`평균 가공중 대기시간`, " \
          "`실 Cycle Time` " \
          "from process_report " \
          "where report_date = %s " \
          "and ent_id = %s " \
          "order by ent_id, machine_no, machine_name, `조업 시작`"
    return yhsdb.__get_yhsdb_query(conn=conn, sql=sql, args=(report_date, ent_id,))

def get_process_report_detail(conn: connect, report_date: str, ent_id: int) -> pd.DataFrame:
    sql = "select %s report_date, " \
          "a.seq, b.id ent_id, a.mkey, a.lot, a.ent ent_code, b.comment ent_name, convert(c.machine_no, decimal) machine_no, a.mid, a.program, a.plan, a.count, a.start, a.end, " \
          "sec_to_time(round(a.period/1000)) 가공시간, " \
          "sec_to_time(round(a.active_time/1000)) 실가공시간, " \
          "sec_to_time(round(a.idle/1000)) 준비교체시간, " \
          "sec_to_time(round((if((a.period-a.active_time) > 0, a.period-a.active_time, a.active_time-a.period))/1000)) 가공중대기시간, " \
          "sec_to_time(round((a.period+a.idle)/1000)) CT, " \
          "mkey.first, " \
          "round(a.idle/1000) `준비교체(초)` " \
          "from process_hist a " \
          "left outer join " \
          "( " \
          "select min(t.seq) first_seq, " \
          "'Y' first, " \
          "t.mkey " \
          "from process_hist t use index (mid_lot) " \
          "where t.ent in (select t1.name from svc_enterprise t1 where t1.id = %s) " \
          "and t.start >= concat(date(%s), ' 00:00:00') " \
          "and t.start <= concat(date(%s), ' 23:59:59') " \
          "group by t.mkey order by mkey " \
          ") mkey " \
          "on a.seq = mkey.first_seq, " \
          "svc_enterprise b, " \
          "svc_cnc c " \
          "where a.start >= concat(date(%s), ' 00:00:00') " \
          "and a.start <= concat(date(%s), ' 23:59:59') " \
          "and b.name = a.ent " \
          "and c.id = a.mkey " \
          "and b.id = %s " \
          "order by b.id, convert(c.machine_no, decimal), a.mid, a.start"
    return yhsdb.__get_yhsdb_query(conn=conn, sql=sql, args=(report_date,ent_id,report_date,report_date,report_date,report_date,ent_id,))

def getDBConnect():
    if env["YHS_DEPLOY_ENV"] == "PRD":
        con = connect(
            unix_socket=env["YHS_DB_DATA_UNIXSOCKET"],  # cloud functions 배포시 host 대신 unix_socket 사용하여 접속
            user=env["YHS_DB_DATA_USER"],
            password=env["YHS_DB_DATA_PW"],
            db=env["YHS_DB_DATA_DB"],
            charset="utf8"
        )
    else:
        con = connect(
            host=env["YHS_DB_DATA_HOST"],
            user=env["YHS_DB_DATA_USER"],
            password=env["YHS_DB_DATA_PW"],
            db=env["YHS_DB_DATA_DB"],
            charset="utf8"
        )
    return con

# 배포된 함수 테스트 방법 : readme.md 참고
def report(event, context) -> (str, int):
    env = dotenv_values(".env")
    try:
        if env["YHS_DEPLOY_ENV"] == "DEV":
            params = event
        else:
            pubsub_message = base64.b64decode(event['data']).decode('utf-8')
            params = json.loads(pubsub_message)
        ent_id = params.get('ent_id')
        report_date = params.get('report_date')
        recv_email_addr = params.get('recv_email_addr')
    except Exception as e:
        log(e)
        return "파라미터가 잘못되었습니다.", 400

    log(' ------ start ------ ')

    db_conn = getDBConnect()

    # ---------------------------------------------------------------------------
    # 사전체크
    # ---------------------------------------------------------------------------
    ent_name: str = get_ent_name(conn=db_conn, ent_id=ent_id)
    if ent_name is None:
        return f"회사 아이디를 찾을수 없습니다. ent_id = {ent_id}", 204
    if recv_email_addr is None:
        recv_email_addr = get_report_recv_emails(conn=db_conn, ent_id=ent_id)
        if recv_email_addr is None:
            return f"메일 수신자가 등록되어 있지 않습니다. ent_id = {ent_id}", 204

    # ---------------------------------------------------------------------------/
    # 리포트를 데이터 프레임 형태로 가져옴
    # ---------------------------------------------------------------------------
    # 기본리포트
    df_report = get_process_report(conn=db_conn, report_date=report_date, ent_id=ent_id)
    # 상세리포트
    df_report_detail = get_process_report_detail(conn=db_conn, report_date=report_date, ent_id=ent_id)
    if len(df_report) == 0 or len(df_report_detail) == 0:
        return f"no report found", 204
    # 휴게시간
    df_report_brktimes = get_process_report_brktimes(conn=db_conn, ent_id=ent_id)

    log(f"{ent_name}(id : {ent_id}) - {report_date} 리포트 생성 시작")

    # ---------------------------------------------------------------------------
    # 상세리포트에 각종 통계 데이터 열 추가
    # ---------------------------------------------------------------------------
    # 'first' 칼럼이 'Y'가 아닌 (null 인) 데이터를 기준으로 'lot' 그룹별 '준비교체(초)'의 평균, 중위수 등을 구해 새로운 열에 추가
    # 'first' 칼럼이 'Y'인 데이터를 제외하는 이유는, 기계별 첫 조업시 준비교체 시간이 크게 잡히기 때문(보통 전일 마감시간부터 계산되어 값이 크므로 통계정보에서 제외)
    df_report_detail['준비교체 평균'] = df_report_detail[df_report_detail['first'].isnull()].groupby('lot')['준비교체(초)'].transform('mean')  # 평균
    df_report_detail['준비교체 중앙값'] = df_report_detail[df_report_detail['first'].isnull()].groupby('lot')['준비교체(초)'].transform('median')  # 중위수
    df_report_detail['준비교체 표준편차'] = df_report_detail[df_report_detail['first'].isnull()].groupby('lot')['준비교체(초)'].transform('std')  # 표준편차
    df_report_detail['준비교체 최소값'] = df_report_detail[df_report_detail['first'].isnull()].groupby('lot')['준비교체(초)'].transform(lambda x: x.astype(float).quantile(0))  # 최소값
    df_report_detail['준비교체 1사분위'] = df_report_detail[df_report_detail['first'].isnull()].groupby('lot')['준비교체(초)'].transform(lambda x: x.astype(float).quantile(0.25))  # 1사분위
    df_report_detail['준비교체 2사분위'] = df_report_detail[df_report_detail['first'].isnull()].groupby('lot')['준비교체(초)'].transform(lambda x: x.astype(float).quantile(0.5))  # 2사분위
    df_report_detail['준비교체 3사분위'] = df_report_detail[df_report_detail['first'].isnull()].groupby('lot')['준비교체(초)'].transform(lambda x: x.astype(float).quantile(0.75))  # 3사분위
    df_report_detail['준비교체 최대값'] = df_report_detail[df_report_detail['first'].isnull()].groupby('lot')['준비교체(초)'].transform(lambda x: x.astype(float).quantile(1))  # 최대값
    df_report_detail['준비교체 IQR'] = df_report_detail['준비교체 3사분위'] - df_report_detail['준비교체 1사분위']
    const_idle_outlier = get_report_const_idle_outlier(conn=db_conn)
    df_report_detail['준비교체 이상치(상한)'] = df_report_detail['준비교체 3사분위'] + (const_idle_outlier * df_report_detail['준비교체 IQR'])

    # ---------------------------------------------------------------------------
    # google storage 에서 리포트 기본폼 로드
    # ---------------------------------------------------------------------------
    if env["YHS_DEPLOY_ENV"] == "DEV":
        workbook = openpyxl.load_workbook("report_form.xlsx")  # google cloud storage 대신 로컬에서 파일을 로드
    else :
        client = storage.Client.from_service_account_json('gc_private-key.json')
        bucket = client.get_bucket('process_report')  # google cloud storage에 생성된 버킷 이름
        blob = bucket.blob('report_form/report_form.xlsx')
        content = blob.download_as_string()
        io_content = io.BytesIO(content)
        workbook = openpyxl.load_workbook(io_content)

    # ---------------------------------------------------------------------------
    # 리포트 제공용 엑셀 폼 오픈
    # ---------------------------------------------------------------------------
    sheet_report = workbook['제공폼']
    sheet_report.title = '리포트'
    sheet_report_detail = workbook['상세제공폼']
    sheet_report_detail.title = '상세리포트'
    sheet_report_brktimes = workbook['휴게시간폼']
    sheet_report_brktimes.title = '휴게시간'

    # ---------------------------------------------------------------------------
    # 엑셀폼에 SQL로 불러온 데이터프레임 추가
    # ---------------------------------------------------------------------------
    log('리포트 붙여넣기 시작')
    for row in dataframe_to_rows(df_report, index=False, header=False):
        sheet_report.append(row)
    for row in dataframe_to_rows(df_report_detail, index=False, header=False):
        sheet_report_detail.append(row)
    for row in dataframe_to_rows(df_report_brktimes, index=False, header=False):
        sheet_report_brktimes.append(row)
    log('리포트 붙여넣기 종료')

    # ---------------------------------------------------------------------------
    # 엑셀에 붙여넣은 데이터도 샘플 데이터와과 동일하게 스타일 적용하기 위한 함수 정의
    # ---------------------------------------------------------------------------
    def getNamedStyleFromCell(cell, style_name: str):
        style = NamedStyle(cell.coordinate + style_name)  # cell.coordinate + style_name 를 적용하지 않으면 스타일 이름 중복 오류가 발생함.
        style.font = copy(cell.font)
        style.fill = copy(cell.fill)
        style.border = copy(cell.border)
        style.alignment = copy(cell.alignment)
        style.number_format = copy(cell.number_format)
        style.protection = copy(cell.protection)
        return style

    def getNamedStylesFromRow(row, style_name: str):
        style_list = []
        for cell in row:
            style_list.append(getNamedStyleFromCell(cell, style_name))
        return style_list

    # 샘플 스타일을 복사
    styles_report = getNamedStylesFromRow(sheet_report[4], "basic")
    styles_report_detail = getNamedStylesFromRow(sheet_report_detail[3], "detail")
    styles_report_brktimes = getNamedStylesFromRow(sheet_report_brktimes[4], "brktimes")

    # ---------------------------------------------------------------------------
    # 기본/상세 리포트 지정범위 서식 적용
    # ---------------------------------------------------------------------------
    # 라인 배경색 서식
    fill_stepping_line = PatternFill(fill_type='solid', start_color='E9F4CC')

    # 로트 구분용 배경색 리스트와 순회참조 이터레이터를 생성(흰색, 회색, 흰색, 회색...)
    fill_stepping_lot_list = itertools.cycle([PatternFill(fill_type='solid', start_color='FFFFFF'),
                                              PatternFill(fill_type='solid', start_color='E8EBEB')])
    fill_stepping_lot = next(fill_stepping_lot_list)
    bef_lot = None

    def adjustStyleToReportRow(row, style_list):
        for i, cell in enumerate(row):
            cell.style = style_list[i]
            # 라인 구분을 위해 홀수라인마다 배경색 적용
            if cell.row % 2 == 1 and 6 <= cell.column <= 24:
                cell.fill = fill_stepping_line

    def adjustStyleToReportDetailRow(row, style_list, bef_lot, fill_stepping_lot):

        marking_cell = None
        idle_sec = -1
        outlier_sec = -1

        for i, cell in enumerate(row):
            cell.style = style_list[i]

            # 라인 구분을 위해 짝수라인마다 배경색 적용
            if cell.row % 2 == 0 and 11 <= cell.column <= 31:
                cell.fill = fill_stepping_line

            # lot 번호체크
            if cell.column == 5:
                cur_lot = cell.value
                # 저장된 lot와 현재 lot가 다를 경우 저장된 lot배경색값을 순회리스트의 다음값으로 교체
                if bef_lot != cur_lot:
                    fill_stepping_lot = next(fill_stepping_lot_list)
                bef_lot = cur_lot  # 현재 lot값 저장
            # lot 구분용 배경색값을 적용
            if 5 <= cell.column <= 10:
                cell.fill = fill_stepping_lot

            # 준비교체시간 셀 저장
            if cell.column == 17:
                marking_cell = cell
            # 준비교체(초) 저장
            elif cell.column == 21 and cell.value is not None:
                idle_sec = float(cell.value)
            # 준비교체(초) 이상치 기준 저장
            elif cell.column == 31 and cell.value is not None:
                outlier_sec = float(cell.value)

        # 준비교체시간 셀에 이상치 기준 초과시 색상 적용
        if (idle_sec > outlier_sec) and (marking_cell is not None) and (outlier_sec > 0):
            marking_cell.font = Font(bold=True, color='FF0000')

        return bef_lot, fill_stepping_lot

    def adjustStyleToBrkTimeRow(row, style_list):
        for i, cell in enumerate(row):
            cell.style = style_list[i]
            # 라인 구분을 위해 홀수라인마다 배경색 적용
            if cell.row % 2 == 1 and 2 <= cell.column <= 3:
                cell.fill = fill_stepping_line

    log('서식 적용 시작')
    # 리포트 : min_row ~ max_row까지, min_col ~ max_col까지)
    for row in sheet_report.iter_rows(min_row=5, max_row=sheet_report.max_row, min_col=1, max_col=sheet_report.max_column):
        adjustStyleToReportRow(row=row, style_list=styles_report)
    # 상세리포트 : min_row ~ max_row까지, min_col ~ max_col까지)
    for row in sheet_report_detail.iter_rows(min_row=4, max_row=sheet_report_detail.max_row, min_col=1, max_col=sheet_report_detail.max_column):
        bef_lot, fill_stepping_lot = adjustStyleToReportDetailRow(row=row, style_list=styles_report_detail, bef_lot=bef_lot, fill_stepping_lot=fill_stepping_lot)
    # 휴게시간 : min_row ~ max_row까지, min_col ~ max_col까지)
    for row in sheet_report_brktimes.iter_rows(min_row=5, max_row=sheet_report_brktimes.max_row, min_col=1, max_col=sheet_report_brktimes.max_column):
        adjustStyleToBrkTimeRow(row=row, style_list=styles_report_brktimes)
    log('서식 적용 종료')

    # ---------------------------------------------------------------------------
    # openpyxl 라이브러로 엑셀 오픈시 기존 메모폼의 크기가 초기화 되므로 재지정 해준다.
    # ---------------------------------------------------------------------------
    for row in sheet_report.iter_rows(min_row=2, max_row=3, min_col=1, max_col=24):
        for cell in row:
            if cell.comment:
                cell.comment.width = 600
                cell.comment.height = 150
    for row in sheet_report_detail.iter_rows(min_row=2, max_row=2, min_col=13, max_col=19):
        for cell in row:
            if cell.comment:
                cell.comment.width = 600
                cell.comment.height = 150

    # -------------------------------------------------------------
    # 엑셀 마무리
    # ---------------------------------------------------------------------------
    # 열숨기기
    sheet_report.column_dimensions['B'].hidden = True
    sheet_report.column_dimensions['C'].hidden = True
    sheet_report.column_dimensions['D'].hidden = True
    sheet_report_detail.column_dimensions['B'].hidden = True
    sheet_report_detail.column_dimensions['C'].hidden = True
    sheet_report_detail.column_dimensions['D'].hidden = True
    sheet_report_detail.column_dimensions['E'].hidden = True
    sheet_report_detail.column_dimensions['F'].hidden = True
    sheet_report_detail.column_dimensions['T'].hidden = True
    sheet_report_detail.column_dimensions['U'].hidden = True
    sheet_report_detail.column_dimensions['V'].hidden = True
    sheet_report_detail.column_dimensions['W'].hidden = True
    sheet_report_detail.column_dimensions['X'].hidden = True
    sheet_report_detail.column_dimensions['Y'].hidden = True
    sheet_report_detail.column_dimensions['Z'].hidden = True
    sheet_report_detail.column_dimensions['AA'].hidden = True
    sheet_report_detail.column_dimensions['AB'].hidden = True
    sheet_report_detail.column_dimensions['AC'].hidden = True
    sheet_report_detail.column_dimensions['AD'].hidden = True
    sheet_report_detail.column_dimensions['AE'].hidden = True

    # 스타일샘플 row 삭제
    sheet_report.delete_rows(4)
    sheet_report.delete_cols(25)
    sheet_report_detail.delete_rows(3)
    sheet_report_brktimes.delete_rows(4)
    # 파일처리
    file = io.BytesIO()
    workbook.save(file)  # 메일 첨부하기 위해 파일을 바이너리로 메모리에 저장

    # -------------------------------------------------------------
    # 메일 발송
    # ---------------------------------------------------------------------------
    log('메일 발송 시작')
    send_email_addr = env['YHS_SEND_EMAIL_ADDR']
    # send_email_pw = env['SEND_EMAIL_PW']
    send_email_pw = env['SEND_EMAIL_APP_PW']  # 앱 패스워드 사용 (구글계정 2단계 보안인증 설정 계정 사용시, 앱 패스워드를 별도 사용하여야 한다. - 설정페이지 https://security.google.com/settings/security/apppasswords )

    smtp = smtplib.SMTP('smtp.gmail.com', 587)
    smtp.ehlo()
    smtp.starttls()  # TLS사용
    smtp.login(send_email_addr, send_email_pw)  # 로그인 에러 발생시 해당계정으로 https://www.google.com/settings/security/lesssecureapps 접속, '보안 수준이 낮은 앱의 액세스 허용'을 활성화 해야함

    # 메일 본문에 붙일 버전 히스토리
    report_ver_hist = get_report_ver_hist(conn=db_conn)

    # 이메일 생성
    msg = MIMEMultipart()
    msg['Subject'] = f'{ent_name} 일일 가공 리포트 ({report_date})'
    msg['From'] = formataddr(('연합시스템', send_email_addr))
    msg['To'] = recv_email_addr
    fileName = f'{report_date} {ent_name} 일일가공리포트.xlsx'
    email_body = f"안녕하세요, 바로 팩토리에서 제공하는 일일 가공 리포트 서비스 메일입니다. \n\n" \
                 f"{fileName} 파일을 첨부하였습니다.\n\n" + report_ver_hist

    msg.attach(MIMEText(email_body, 'plain'))

    # Excel 파일 첨부
    attachment = MIMEApplication(file.getvalue(), _subtype='xlsx')
    attachment.add_header('Content-Disposition', 'attachment', filename=fileName)
    msg.attach(attachment)

    # 이메일 전송
    smtp.sendmail(send_email_addr, recv_email_addr.split(', '), msg.as_string())

    # 이메일 서버 연결 종료
    smtp.quit()
    log('메일 발송 종료')

    # ---------------------------------------------------------------------------
    # 작업 마무리
    # ---------------------------------------------------------------------------
    db_conn.close()
    log(' ------ end ------ ')
    return ' ------ Success ------ ', 200

# 배포된 함수 테스트 방법 : readme.md 참고
def call_report(event, context):
    try:
        log(f'event: {event}')
        pubsub_message = base64.b64decode(event['data']).decode('utf-8')
        log(f'pubsub_message: {pubsub_message}')
        params = json.loads(pubsub_message)
        log(f'params: {params}')
        report_date = params.get('report_date')
        only_receive_email = params.get('only_receive_email')
    except Exception as e:
        log(f'Exception: {e}')
        return "파라미터가 잘못되었습니다.", 400

    log(' ------ start ------ ')

    env = dotenv_values(".env")
    topic_path = env["YHS_PUBSUB_TOPIC"]

    # ---------------------------------------------------------------------------
    # 사전체크
    if report_date == "0000-00-00":
        report_date = get_daystr(day_offset=-1)

    # -------------------------------------------------------------
    # GCP Service Account Key File Path 글로벌 환경변수로 설정
    #  - os.environ으로 설정하는 환경변수는 현재 프로세스 및 자식 프로세스에만 적용된다.
    #    (여러 세션이나 재부팅에 걸쳐 지속되지 않으며 프로세스가 다시 시작되면 재설정 해야 한다.)
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = './gc_private-key.json'
    # -------------------------------------------------------------

    # -------------------------------------------------------------
    # DB Connection
    con = getDBConnect()
    df = get_ent_ids(conn=con)
    ent_id_list = df['ent_id'].tolist()
    # -------------------------------------------------------------

    publisher = pubsub_v1.PublisherClient()

    log(f'{report_date} 일자 리포트 발행, 대상 업체 수: {len(ent_id_list)}')

    for ent_id in ent_id_list:

        # Publish message to topic
        # only_receive_email 파라미터가 있으면 해당 이메일로만 발송
        if only_receive_email is not None:
            message = json.dumps({"ent_id": ent_id, "report_date": report_date, "recv_email_addr": only_receive_email})
        else:
            message = json.dumps({"ent_id": ent_id, "report_date": report_date})
        message_bytes = message.encode('utf-8')
        future = publisher.publish(topic_path, data=message_bytes)
        message_id = future.result()

        log(f'Message published with ID: {message_id}')

    log(' ------  end  main ------ ')

    return "success", 200

def get_ent_ids(conn: connect) -> pd.DataFrame:
    sql = "select convert(cd, UNSIGNED) ent_id, cd_name ent_name from ref_code where grp_cd = 'proc_report_recvs' and value_str is not null order by convert(cd, UNSIGNED)"
    return yhsdb.__get_yhsdb_query(conn=conn, sql=sql)


if __name__ == '__main__':
    # 로컬 테스트 코드
    if env["YHS_DEPLOY_ENV"] == "DEV":
        # report(event={"ent_id": "1", "report_date": "2023-05-24", "recv_email_addr": "Hwang.Daeyeon@yhsbearing.com, allther@yhsbearing.com"}, context=None)
        result = report(event={"ent_id": "1", "report_date": "2023-06-01", "recv_email_addr": "Hwang.Daeyeon@yhsbearing.com"}, context=None)
        print(result)
